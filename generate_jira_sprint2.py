from __future__ import annotations

import json
import os
import re
import shutil
import time
from pathlib import Path
from typing import Any

from dotenv import load_dotenv
from openai import OpenAI
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

DATASET_DIR = Path(__file__).resolve().parent
load_dotenv(DATASET_DIR / ".env")

MODEL = os.getenv("NUTRIVANA_OPENAI_MODEL", "openai/gpt-4o-mini")

SPRINT_NAME = "Sprint 2"
OUT_PATH = DATASET_DIR / "generated" / "jira" / "sprint2_jira.xlsx"
BLUEPRINT_PATH = DATASET_DIR / "sprint_2.md"
EXPECTED_TICKETS = 6

SYSTEM_PROMPT = (
    "You are writing realistic Jira tickets for a startup called Nutrivana which is a "
    "nutrition tracking app for health conscious Indians. The team members are Shristi "
    "Sharmistha (CEO/PM), Arjun Mehta (CTO), Priya Nair (Frontend Engineer), Kabir Sharma "
    "(Designer), Ananya Iyer (Marketing). Write a complete realistic Jira ticket based on "
    "the blueprint provided. Do not copy the blueprint word for word. Use it as "
    "instructions to write natural realistic workplace content."
)


def _client() -> OpenAI:
    key = os.getenv("OPENAI_API_KEY", "").strip()
    base_url = os.getenv("OPENAI_BASE_URL", "").strip() or None
    if not key or key == "sk-or-v1-your-key-here":
        raise RuntimeError("Set OPENAI_API_KEY in nutrivana-dataset/.env")
    kwargs: dict[str, Any] = {"api_key": key}
    if base_url:
        kwargs["base_url"] = base_url
    return OpenAI(**kwargs)


def _tier_default_min_comments(tier: int) -> int:
    if tier == 1:
        return 8
    if tier == 2:
        return 4
    return 2


# Optional per-ticket overrides (same pattern as Sprint 1; empty for Sprint 2 defaults).
_MIN_COMMENTS_BY_TICKET: dict[str, int] = {}


def _required_min_comments(ticket_key: str, tier: int) -> int:
    return _MIN_COMMENTS_BY_TICKET.get(ticket_key, _tier_default_min_comments(tier))


def _comment_rule_phrase(min_comments: int) -> str:
    return f"at least {min_comments} comments in the comments array (hard requirement)"


def _split_ticket_blocks(markdown: str) -> list[str]:
    blocks = re.split(r"\n(?=### Ticket \d+:\s*[A-Z0-9-]+\s*$)", markdown, flags=re.MULTILINE)
    out: list[str] = []
    for b in blocks:
        b = b.strip()
        if b.startswith("### Ticket "):
            out.append(b)
    return out


def _parse_ticket_key(block: str) -> str:
    m = re.match(r"^### Ticket \d+:\s*([A-Z0-9-]+)\s*$", block, re.MULTILINE)
    if not m:
        raise ValueError(f"Could not parse ticket key from block start: {block[:80]!r}")
    return m.group(1).strip()


def _parse_tier(block: str) -> int:
    m = re.search(r"\*\*Tier:\*\*\s*(\d+)", block)
    if not m:
        raise ValueError(f"Could not parse Tier in ticket {_parse_ticket_key(block)}")
    return int(m.group(1))


def _valid_comment_entries(comments: Any) -> list[dict[str, Any]]:
    if not isinstance(comments, list):
        return []
    out: list[dict[str, Any]] = []
    for c in comments:
        if not isinstance(c, dict):
            continue
        text = str(c.get("text", "")).strip()
        if not text:
            continue
        out.append(c)
    return out


def _call_ticket_json(
    client: OpenAI,
    *,
    ticket_key: str,
    tier: int,
    blueprint_block: str,
    max_rounds: int = 12,
) -> dict[str, Any]:
    min_comments = _required_min_comments(ticket_key, tier)
    comment_rule = _comment_rule_phrase(min_comments)
    user = (
        f"You are generating ONE Jira ticket for ticket ID **{ticket_key}**.\n\n"
        f"HARD REQUIREMENT: the comments array MUST contain {comment_rule}. "
        f"Fewer than {min_comments} comments will be rejected.\n\n"
        "Return ONLY valid JSON (no markdown, no code fences).\n\n"
        "JSON shape:\n"
        "{\n"
        '  "ticket_id": "string (must match blueprint)",\n'
        '  "title": "string",\n'
        '  "type": "Epic|Story|Task|Bug",\n'
        '  "status": "To Do|In Progress|Done",\n'
        '  "priority": "P0|P1|P2",\n'
        '  "assignee": "full name as in blueprint",\n'
        f'  "sprint": "{SPRINT_NAME}",\n'
        '  "story_points": number,\n'
        '  "description": "3-5 sentences, natural workplace language",\n'
        '  "acceptance_criteria": ["bullet 1", "bullet 2", ...],\n'
        '  "comments": [\n'
        "    {\n"
        '      "comment_number": 1,\n'
        '      "author": "full name",\n'
        '      "date": "use realistic January 2025 dates consistent with blueprint",\n'
        '      "text": "full comment body; conversational; not a summary"\n'
        "    }\n"
        "  ]\n"
        "}\n\n"
        f"Comments: {comment_rule}. Each comment must read like a real thread: "
        "questions, concerns, updates, decisions. Use Nutrivana-specific terms (EER, USDA, "
        "Supabase RLS, food diary, DV%, macros/micros, React) where appropriate. "
        "Comments must NOT be one-line summaries.\n\n"
        "Map blueprint **Type** to Jira type: Epic stays Epic; Technical task -> Task; "
        "Spike -> Story.\n"
        "Map blueprint **Status** Closed -> Done; In Progress -> In Progress.\n\n"
        "Blueprint for this ticket:\n"
        "-----\n"
        f"{blueprint_block}\n"
        "-----\n"
    )

    messages: list[dict[str, str]] = [
        {"role": "system", "content": SYSTEM_PROMPT},
        {"role": "user", "content": user},
    ]
    last_err: Exception | None = None
    for attempt in range(max_rounds):
        try:
            resp = client.chat.completions.create(
                model=MODEL,
                messages=messages,
                temperature=0.35,
                response_format={"type": "json_object"},
            )
            text = (resp.choices[0].message.content or "").strip()
            data = json.loads(text)
            if str(data.get("ticket_id", "")).strip() != ticket_key:
                raise ValueError(
                    f"ticket_id must be {ticket_key!r}, got {data.get('ticket_id')!r}"
                )
            comments = _valid_comment_entries(data.get("comments"))
            if len(comments) < min_comments:
                raise ValueError(
                    f"Need at least {min_comments} non-empty comments; "
                    f"got {len(comments)} valid (raw array length {len(data.get('comments') or [])})"
                )
            data["comments"] = comments
            return data
        except Exception as e:  # noqa: BLE001
            last_err = e
            messages.append(
                {
                    "role": "user",
                    "content": (
                        "Your previous JSON was invalid or did not meet requirements. "
                        f"Return ONLY valid JSON. Ensure ticket_id is {ticket_key!r}. "
                        f"The comments array MUST have at least {min_comments} objects, each "
                        "with non-empty text. Do not use empty strings as placeholders. "
                        f"Error: {e}"
                    ),
                }
            )
            time.sleep(min(2.0, 0.5 * (attempt + 1)))

    raise RuntimeError(f"Failed after {max_rounds} rounds: {last_err}")


def _autosize(ws) -> None:
    for col in range(1, ws.max_column + 1):
        max_len = 0
        for row in range(1, ws.max_row + 1):
            v = ws.cell(row=row, column=col).value
            if v is None:
                continue
            max_len = max(max_len, min(len(str(v)), 200))
        ws.column_dimensions[get_column_letter(col)].width = min(max(12, max_len + 2), 70)


def _write_workbook(rows_tickets: list[list[Any]], rows_comments: list[list[Any]]) -> Path:
    OUT_PATH.parent.mkdir(parents=True, exist_ok=True)
    wb = Workbook()
    ws_t = wb.active
    ws_t.title = "Tickets"
    headers_t = [
        "Ticket ID",
        "Title",
        "Type",
        "Status",
        "Priority",
        "Assignee",
        "Sprint",
        "Story Points",
        "Description",
        "Acceptance Criteria",
    ]
    ws_t.append(headers_t)
    fill = PatternFill("solid", fgColor="1F4E79")
    font = Font(color="FFFFFF", bold=True)
    for c in range(1, len(headers_t) + 1):
        cell = ws_t.cell(row=1, column=c)
        cell.fill = fill
        cell.font = font
        cell.alignment = Alignment(wrap_text=True, vertical="top")
    for r in rows_tickets:
        ws_t.append(r)
    for row in ws_t.iter_rows(min_row=2, max_row=ws_t.max_row, min_col=1, max_col=ws_t.max_column):
        for cell in row:
            cell.alignment = Alignment(wrap_text=True, vertical="top")
    _autosize(ws_t)

    ws_c = wb.create_sheet(title="Comments")
    headers_c = ["Ticket ID", "Comment Number", "Author", "Date", "Comment Text"]
    ws_c.append(headers_c)
    for c in range(1, len(headers_c) + 1):
        cell = ws_c.cell(row=1, column=c)
        cell.fill = fill
        cell.font = font
        cell.alignment = Alignment(wrap_text=True, vertical="top")
    for r in rows_comments:
        ws_c.append(r)
    for row in ws_c.iter_rows(min_row=2, max_row=ws_c.max_row, min_col=1, max_col=ws_c.max_column):
        for cell in row:
            cell.alignment = Alignment(wrap_text=True, vertical="top")
    _autosize(ws_c)

    tmp_path = OUT_PATH.with_suffix(".tmp.xlsx")
    wb.save(tmp_path)
    try:
        os.replace(tmp_path, OUT_PATH)
        return OUT_PATH
    except PermissionError:
        fallback = OUT_PATH.with_name("sprint2_jira_ready.xlsx")
        shutil.move(str(tmp_path), str(fallback))
        return fallback


def main() -> int:
    md = BLUEPRINT_PATH.read_text(encoding="utf-8")
    blocks = _split_ticket_blocks(md)
    if len(blocks) != EXPECTED_TICKETS:
        raise RuntimeError(f"Expected {EXPECTED_TICKETS} ticket blocks, found {len(blocks)}")

    client = _client()
    rows_tickets: list[list[Any]] = []
    rows_comments: list[list[Any]] = []

    for block in blocks:
        key = _parse_ticket_key(block)
        tier = _parse_tier(block)
        need = _required_min_comments(key, tier)
        print(f"Generating {key} (blueprint Tier {tier}, min comments {need})...")
        data = _call_ticket_json(client, ticket_key=key, tier=tier, blueprint_block=block)

        tid = str(data.get("ticket_id", key)).strip()
        ac = data.get("acceptance_criteria") or []
        ac_text = "\n".join(f"• {str(x).strip()}" for x in ac if str(x).strip())

        rows_tickets.append(
            [
                tid,
                str(data.get("title", "")).strip(),
                str(data.get("type", "")).strip(),
                str(data.get("status", "")).strip(),
                str(data.get("priority", "")).strip(),
                str(data.get("assignee", "")).strip(),
                str(data.get("sprint", SPRINT_NAME)).strip(),
                data.get("story_points", ""),
                str(data.get("description", "")).strip(),
                ac_text,
            ]
        )

        comments = sorted(
            (data.get("comments") or []),
            key=lambda c: int(c.get("comment_number", 0) or 0),
        )
        for c in comments:
            rows_comments.append(
                [
                    tid,
                    c.get("comment_number", ""),
                    str(c.get("author", "")).strip(),
                    str(c.get("date", "")).strip(),
                    str(c.get("text", "")).strip(),
                ]
            )

    saved = _write_workbook(rows_tickets, rows_comments)
    if saved != OUT_PATH:
        print(
            f"\nSaved to {saved} (could not overwrite {OUT_PATH} — file may be open in Excel). "
            "Close sprint2_jira.xlsx and rename/copy the new file, or rerun this script."
        )
    else:
        print(f"\nWrote {saved}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
