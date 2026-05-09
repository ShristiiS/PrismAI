from __future__ import annotations

import json
import os
import re
import shutil
import sys
import time
from pathlib import Path
from typing import Any

from dotenv import load_dotenv
from openai import OpenAI
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

DATASET_DIR = Path(__file__).resolve().parent
load_dotenv(DATASET_DIR / ".env")

MODEL = os.getenv("NUTRIVANA_OPENAI_MODEL", "openai/gpt-4o-mini")

# User referenced sprint_10-13_planning.md; in this repo the file is Sprint_10-13_Planning.md.
SOURCE_MARKDOWN = DATASET_DIR / "Sprint_10-13_Planning.md"
OUT_PATH = DATASET_DIR / "generated" / "excel" / "sprint10_retro.xlsx"

SYSTEM_PROMPT = """ROLE: You are Shristi Sharmistha writing a sprint retrospective for Nutrivana after the sprint has ended.

CONTEXT: Nutrivana is a nutrition tracking app for health-conscious Indians. The team is Shristi (PM), Arjun (CTO), Priya (Frontend), Kabir (Designer), Ananya (Marketing).

INSTRUCTION: Read the sprint retrospective blueprint carefully and write a realistic honest retrospective. What went well must celebrate real specific wins. What did not go well must name real problems directly without blame. Key learnings must be actionable. Action items must have specific owners and due dates.

PERFORMANCE:
- Every point must reference a specific ticket or decision from the sprint
- What went well must feel genuinely proud not generic
- What did not go well must be honest not defensive
- Action items must be specific enough that a team member can act on them immediately
- Avoid generic statements like 'improve communication' or 'do better next time'
- Team health rating must feel real — not always 5 out of 5

SPECIFICATION:
- Return JSON with keys: summary, delivery_table, went_well, did_not_go_well, key_learnings, decisions, action_items, team_health_rating, team_health_note, metrics
- Each section must have specific content from the blueprint
- metrics must be a JSON array of objects; each object must have ONLY the keys metric and value (both non-empty strings)."""


class GenerationError(RuntimeError):
    pass


_TICKET_ID_RE = re.compile(r"\b[A-Z]{2,}(?:-[A-Z0-9]+)*-\d+\b")


def _sentence_count(text: str) -> int:
    t = text.strip()
    if not t:
        return 0
    parts = re.split(r"(?<=[.!?])\s+", t)
    return len([p for p in parts if p.strip()])


def _extract_sprint10_retro_blueprint(full_md: str) -> str:
    start_tag = "# SPRINT 10 RETROSPECTIVE"
    start = full_md.find(start_tag)
    if start == -1:
        raise GenerationError(f"Could not find {start_tag!r} in {SOURCE_MARKDOWN.name}")
    # Sprint 11 follows Sprint 10 in this file.
    end_tag = "# SPRINT 11 PLANNING DOCUMENT"
    end = full_md.find(end_tag, start)
    if end == -1:
        # Fallback: stop at end of file
        end = len(full_md)
    return full_md[start:end].strip()


def _client() -> OpenAI:
    key = os.getenv("OPENAI_API_KEY", "").strip()
    base_url = os.getenv("OPENAI_BASE_URL", "").strip() or None
    if not key or key == "sk-or-v1-your-key-here":
        raise GenerationError(
            "OpenAI API key missing. Set OPENAI_API_KEY in nutrivana-dataset/.env (use python-dotenv)."
        )
    kwargs: dict[str, Any] = {"api_key": key}
    if base_url:
        kwargs["base_url"] = base_url
    return OpenAI(**kwargs)


def _validate_narrative_bullets(items: Any, field: str, *, min_items: int, max_items: int) -> None:
    if not isinstance(items, list):
        raise ValueError(f"{field} must be an array")
    n = len(items)
    if not (min_items <= n <= max_items):
        raise ValueError(f"{field} must have between {min_items} and {max_items} items, got {n}")
    for i, item in enumerate(items):
        if not isinstance(item, str):
            raise ValueError(f"{field}[{i}] must be a string")
        t = item.strip()
        if _sentence_count(t) < 2:
            raise ValueError(f"{field}[{i}] must have at least 2 sentences")
        if not _TICKET_ID_RE.search(t):
            raise ValueError(f"{field}[{i}] must reference at least one ticket ID (e.g. MON-001)")


def _validate_spec(data: dict[str, Any]) -> None:
    if not str(data.get("summary", "")).strip():
        raise ValueError("summary must be non-empty string")
    dt = data.get("delivery_table")
    if not isinstance(dt, list) or not dt:
        raise ValueError("delivery_table must be a non-empty array")
    for i, row in enumerate(dt):
        if not isinstance(row, dict):
            raise ValueError(f"delivery_table[{i}] must be an object")
        for k in ("ticket_id", "committed", "delivered", "notes"):
            if k not in row:
                raise ValueError(f"delivery_table[{i}] missing {k}")

    _validate_narrative_bullets(data.get("went_well"), "went_well", min_items=3, max_items=8)
    _validate_narrative_bullets(
        data.get("did_not_go_well"), "did_not_go_well", min_items=2, max_items=8
    )

    kl = data.get("key_learnings")
    if not isinstance(kl, list) or not (3 <= len(kl) <= 6):
        raise ValueError("key_learnings must be 3-6 items")

    decisions = data.get("decisions")
    if not isinstance(decisions, list) or not decisions:
        raise ValueError("decisions must be a non-empty array")

    actions = data.get("action_items")
    if not isinstance(actions, list) or not actions:
        raise ValueError("action_items must be a non-empty array")
    for i, a in enumerate(actions):
        if not isinstance(a, dict):
            raise ValueError(f"action_items[{i}] must be an object")
        for k in ("action", "owner", "due_date"):
            if not str(a.get(k, "")).strip():
                raise ValueError(f"action_items[{i}].{k} must be non-empty")

    rating = data.get("team_health_rating")
    if not (
        isinstance(rating, (int, float))
        or (isinstance(rating, str) and rating.strip() and rating.strip()[0].isdigit())
    ):
        raise ValueError("team_health_rating must be 1-5 (number) or like '4/5'")
    if not str(data.get("team_health_note", "")).strip():
        raise ValueError("team_health_note must be non-empty")

    metrics = data.get("metrics")
    if not isinstance(metrics, list) or not metrics:
        raise ValueError("metrics must be a non-empty array")
    for i, m in enumerate(metrics):
        if not isinstance(m, dict):
            raise ValueError(f"metrics[{i}] must be an object")
        if set(m.keys()) != {"metric", "value"}:
            raise ValueError(f"metrics[{i}] must have ONLY keys metric and value")
        if m.get("metric") is None or not str(m.get("metric")).strip():
            raise ValueError(f"metrics[{i}].metric empty")
        if m.get("value") is None or not str(m.get("value")).strip():
            raise ValueError(f"metrics[{i}].value empty")


def _call_retro_json(client: OpenAI, *, blueprint: str) -> dict[str, Any]:
    user = (
        "Return ONLY valid JSON (no markdown, no code fences) for the Sprint 10 retrospective.\n\n"
        "JSON schema:\n"
        "{\n"
        '  "summary": "string (overall what happened)",\n'
        '  "delivery_table": [{"ticket_id":"string","committed":"Yes/No","delivered":"Yes/No","notes":"string"}],\n'
        '  "went_well": ["string (>=2 sentences, must include ticket id)", "..."],\n'
        '  "did_not_go_well": ["string (>=2 sentences, must include ticket id)", "..."],\n'
        '  "key_learnings": ["string", "..."],\n'
        '  "decisions": ["string", "..."],\n'
        '  "action_items": [{"action":"string","owner":"string","due_date":"string"}],\n'
        '  "team_health_rating": "4/5 or 1-5",\n'
        '  "team_health_note": "string",\n'
        '  "metrics": [{"metric":"string","value":"string"}]\n'
        "}\n\n"
        "Hard requirements:\n"
        "- Metrics must be objects with metric+value only; BOTH fields non-empty.\n"
        "- went_well must have at least 3 bullets; did_not_go_well at least 2; each bullet >= 2 sentences and contains a ticket id.\n\n"
        "Blueprint:\n-----\n"
        f"{blueprint}\n"
        "-----\n"
    )

    messages: list[dict[str, str]] = [
        {"role": "system", "content": SYSTEM_PROMPT},
        {"role": "user", "content": user},
    ]

    max_attempts = 4  # 1 initial + 3 retries
    for attempt in range(max_attempts):
        try:
            resp = client.chat.completions.create(
                model=MODEL,
                messages=messages,
                temperature=0.35,
                response_format={"type": "json_object"},
            )
            text = (resp.choices[0].message.content or "").strip()
            data = json.loads(text)
            _validate_spec(data)
            return data
        except Exception as e:  # noqa: BLE001
            if attempt == max_attempts - 1:
                msg = f"Validation failed after 3 retries: {e}"
                print(msg, file=sys.stderr)
                raise GenerationError(msg) from e
            messages.append(
                {
                    "role": "user",
                    "content": (
                        "Your previous JSON failed validation. Return ONLY valid JSON.\n"
                        "Fix metrics so every row has BOTH metric and value non-empty.\n"
                        "Ensure went_well and did_not_go_well bullets are >=2 sentences and include ticket IDs.\n"
                        f"Validation error: {e}"
                    ),
                }
            )
            time.sleep(min(2.0, 0.5 * (attempt + 1)))
    raise GenerationError("Unreachable")


def _autosize_columns(ws) -> None:
    for col in range(1, ws.max_column + 1):
        max_len = 0
        for row in range(1, ws.max_row + 1):
            v = ws.cell(row=row, column=col).value
            if v is None:
                continue
            max_len = max(max_len, min(len(str(v)), 200))
        ws.column_dimensions[get_column_letter(col)].width = min(max(12, max_len + 2), 70)


def _style_header_row(ws, col_count: int) -> None:
    fill = PatternFill("solid", fgColor="1F4E79")
    font = Font(color="FFFFFF", bold=True)
    for c in range(1, col_count + 1):
        cell = ws.cell(row=1, column=c)
        cell.fill = fill
        cell.font = font
        cell.alignment = Alignment(wrap_text=True, vertical="top")


def _write_workbook(spec: dict[str, Any], out_path: Path) -> Path:
    wb = Workbook()
    wb.remove(wb.active)

    # Sheet 1 — Sprint Summary
    ws1 = wb.create_sheet(title="Sprint Summary")
    ws1.freeze_panes = "A2"
    ws1.append(["Ticket", "Committed", "Delivered", "Notes"])
    _style_header_row(ws1, 4)
    for row in spec["delivery_table"]:
        ws1.append(
            [
                str(row.get("ticket_id", "")).strip(),
                str(row.get("committed", "")).strip(),
                str(row.get("delivered", "")).strip(),
                str(row.get("notes", "")).strip(),
            ]
        )
    for row in ws1.iter_rows(min_row=1, max_row=ws1.max_row, min_col=1, max_col=4):
        for cell in row:
            cell.alignment = Alignment(wrap_text=True, vertical="top")
    _autosize_columns(ws1)

    # Sheet 2 — Retrospective Narrative
    ws2 = wb.create_sheet(title="Retrospective Narrative")
    ws2.freeze_panes = "A2"
    ws2.column_dimensions["A"].width = 28
    ws2.column_dimensions["B"].width = 80
    ws2.column_dimensions["C"].width = 40
    ws2.cell(row=1, column=1, value="Retrospective Narrative").font = Font(bold=True, size=14)
    ws2.merge_cells(start_row=1, start_column=1, end_row=1, end_column=3)

    def add_section(title: str, body: str) -> None:
        r = ws2.max_row + 1
        ws2.cell(row=r, column=1, value=title).font = Font(bold=True)
        ws2.cell(row=r, column=2, value=body).alignment = Alignment(wrap_text=True, vertical="top")

    def numbered(items: list[Any]) -> str:
        return "\n".join([f"{i}. {str(x).strip()}" for i, x in enumerate(items, start=1)])

    add_section("Sprint Summary", str(spec["summary"]).strip())
    add_section("What Went Well", numbered(list(spec["went_well"])))
    add_section("What Did Not Go Well", numbered(list(spec["did_not_go_well"])))
    add_section("Key Learnings", numbered(list(spec["key_learnings"])))
    add_section("Decisions Made", numbered(list(spec["decisions"])))

    r = ws2.max_row + 1
    ws2.cell(row=r, column=1, value="Action Items").font = Font(bold=True)
    r += 1
    for c, h in enumerate(["Action", "Owner", "Due Date"], start=1):
        cell = ws2.cell(row=r, column=c, value=h)
        cell.fill = PatternFill("solid", fgColor="1F4E79")
        cell.font = Font(color="FFFFFF", bold=True)
        cell.alignment = Alignment(wrap_text=True, vertical="top")
    for a in spec["action_items"]:
        r += 1
        ws2.cell(row=r, column=1, value=str(a["action"]).strip()).alignment = Alignment(
            wrap_text=True, vertical="top"
        )
        ws2.cell(row=r, column=2, value=str(a["owner"]).strip()).alignment = Alignment(
            wrap_text=True, vertical="top"
        )
        ws2.cell(row=r, column=3, value=str(a["due_date"]).strip()).alignment = Alignment(
            wrap_text=True, vertical="top"
        )

    r += 1
    ws2.cell(row=r, column=1, value="Team Health").font = Font(bold=True)
    ws2.cell(row=r, column=2, value=spec["team_health_rating"])
    ws2.cell(row=r, column=3, value=str(spec["team_health_note"]).strip()).alignment = Alignment(
        wrap_text=True, vertical="top"
    )

    # Sheet 3 — Metrics This Sprint
    ws3 = wb.create_sheet(title="Metrics This Sprint")
    ws3.freeze_panes = "A2"
    ws3.append(["Metric", "Value"])
    _style_header_row(ws3, 2)
    for m in spec["metrics"]:
        ws3.append([str(m["metric"]).strip(), str(m["value"]).strip()])
    for row in ws3.iter_rows(min_row=2, max_row=ws3.max_row, min_col=1, max_col=2):
        for cell in row:
            cell.alignment = Alignment(wrap_text=True, vertical="top")
    _autosize_columns(ws3)

    out_path.parent.mkdir(parents=True, exist_ok=True)
    tmp = out_path.with_suffix(".tmp.xlsx")
    wb.save(tmp)
    try:
        os.replace(tmp, out_path)
        return out_path
    except PermissionError:
        fb = out_path.with_name("sprint10_retro_ready.xlsx")
        shutil.move(str(tmp), str(fb))
        return fb


def main() -> int:
    full_md = SOURCE_MARKDOWN.read_text(encoding="utf-8")
    blueprint = _extract_sprint10_retro_blueprint(full_md)

    print(f"Model: {MODEL}")
    print("Calling API for Sprint 10 retrospective...")
    client = _client()
    spec = _call_retro_json(client, blueprint=blueprint)
    saved = _write_workbook(spec, OUT_PATH)
    if saved != OUT_PATH:
        print(
            f"Saved to {saved} (could not overwrite {OUT_PATH} — file may be open in Excel). "
            "Close sprint10_retro.xlsx and rename/copy the new file, or rerun."
        )
    else:
        print(f"Wrote {saved}")
    return 0


if __name__ == "__main__":
    try:
        raise SystemExit(main())
    except GenerationError as e:
        print(str(e), file=sys.stderr)
        raise SystemExit(1)

