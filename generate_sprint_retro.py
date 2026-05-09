from __future__ import annotations

import argparse
import json
import os
import re
import shutil
import sys
import time
from pathlib import Path
from typing import Any

from dotenv import load_dotenv
from openai import OpenAI
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

DATASET_DIR = Path(__file__).resolve().parent
load_dotenv(DATASET_DIR / ".env")

MODEL = os.getenv("NUTRIVANA_OPENAI_MODEL", "openai/gpt-4o-mini")

SOURCE_MARKDOWN = DATASET_DIR / "Sprint_Planning_and_Ret.md"
OUT_EXCEL_DIR = DATASET_DIR / "generated" / "excel"

SYSTEM_PROMPT = """ROLE: You are Shristi Sharmistha writing a sprint retrospective for Nutrivana after the sprint has ended.

CONTEXT: Nutrivana is a nutrition tracking app for health-conscious Indians. The team is Shristi (PM), Arjun (CTO), Priya (Frontend), Kabir (Designer), Ananya (Marketing).

INSTRUCTION: Read the sprint retrospective blueprint carefully and write a realistic honest retrospective. What went well must celebrate real specific wins. What did not go well must name real problems directly without blame. Key learnings must be actionable. Action items must have specific owners and due dates.

PERFORMANCE:
- Every point must reference a specific ticket or decision from the sprint
- What went well must feel genuinely proud not generic
- What did not go well must be honest not defensive
- Action items must be specific enough that a team member can act on them immediately
- Avoid generic statements like 'improve communication' or 'do better next time'
- Team health rating must feel real — not always 5 out of 5

SPECIFICATION:
- Return JSON with keys: summary, delivery_table, went_well, did_not_go_well, key_learnings, decisions, action_items, team_health_rating, team_health_note, metrics
- Each section must have specific content from the blueprint
- metrics must be a JSON array of objects; each object must have ONLY the keys metric and value (both non-empty strings)."""


class GenerationError(RuntimeError):
    pass


# Matches Nutrivana-style ticket keys (e.g. TECH-001, GOAL-SPIKE-001, NUTR-TASK-003, AN-CR-001).
_TICKET_ID_RE = re.compile(r"\b[A-Z]{2,}(?:-[A-Z0-9]+)*-\d+\b")


def _sentence_count(text: str) -> int:
    """Rough sentence count for validation (splits on . ? ! followed by whitespace)."""
    t = text.strip()
    if not t:
        return 0
    parts = re.split(r"(?<=[.!?])\s+", t)
    return len([p for p in parts if p.strip()])


def _extract_retro_blueprint(full_md: str, sprint_number: int) -> str:
    start_tag = f"## SPRINT {sprint_number} RETROSPECTIVE"
    start = full_md.find(start_tag)
    if start == -1:
        raise GenerationError(
            f"Could not find {start_tag!r} in Sprint_Planning_and_Ret.md"
        )
    if sprint_number < 9:
        end_tag = f"\n# SPRINT {sprint_number + 1}\n"
        end = full_md.find(end_tag, start)
        if end == -1:
            raise GenerationError(
                f"Could not find end marker {end_tag!r} after sprint {sprint_number} retrospective"
            )
    else:
        end_marker = "\n## SUMMARY OF ALL SPRINT DOCUMENTS"
        end = full_md.find(end_marker, start)
        if end == -1:
            end = len(full_md)
    return full_md[start:end].strip()


def _retro_out_path(sprint_number: int) -> Path:
    return OUT_EXCEL_DIR / f"sprint{sprint_number}_retro.xlsx"


def _client() -> OpenAI:
    key = os.getenv("OPENAI_API_KEY", "").strip()
    base_url = os.getenv("OPENAI_BASE_URL", "").strip() or None
    if not key or key == "sk-or-v1-your-key-here":
        raise GenerationError(
            "OpenAI API key missing. Set OPENAI_API_KEY in nutrivana-dataset/.env (use python-dotenv)."
        )
    kwargs: dict[str, Any] = {"api_key": key}
    if base_url:
        kwargs["base_url"] = base_url
    return OpenAI(**kwargs)


def _call_retro_json(client: OpenAI, *, user: str) -> dict[str, Any]:
    """Up to 4 API calls (initial + 3 retries) until JSON parses and passes strict validation."""
    messages: list[dict[str, str]] = [
        {"role": "system", "content": SYSTEM_PROMPT},
        {"role": "user", "content": user},
    ]
    max_attempts = 4  # 1 initial + 3 retries
    for attempt in range(max_attempts):
        try:
            resp = client.chat.completions.create(
                model=MODEL,
                messages=messages,
                temperature=0.35,
                response_format={"type": "json_object"},
            )
            text = (resp.choices[0].message.content or "").strip()
            data = json.loads(text)
            _validate_retro_spec(data)
            return data
        except Exception as e:  # noqa: BLE001
            if attempt == max_attempts - 1:
                msg = f"Validation failed after 3 retries: {e}"
                print(msg, file=sys.stderr)
                raise GenerationError(msg) from e
            messages.append(
                {
                    "role": "user",
                    "content": (
                        "Your previous JSON failed validation. Return ONLY valid JSON.\n"
                        "Fix ALL of the following if applicable:\n"
                        "- metrics: MUST be a non-empty array of objects. Each object MUST have EXACTLY two keys: "
                        '"metric" and "value" (both non-empty strings). No other keys. No string elements.\n'
                        "- went_well: at least 3 strings; each string at least 2 full sentences; each must include "
                        "a ticket ID matching patterns like TECH-001 or GOAL-SPIKE-001.\n"
                        "- did_not_go_well: at least 2 strings; each string at least 2 full sentences; each must "
                        "include a ticket ID as above.\n"
                        f"Validation error was: {e}"
                    ),
                }
            )
            time.sleep(min(2.0, 0.5 * (attempt + 1)))


def _validate_retro_spec(data: dict[str, Any]) -> None:
    if not str(data.get("summary", "")).strip():
        raise ValueError("summary must be non-empty string")
    dt = data.get("delivery_table")
    if not isinstance(dt, list) or len(dt) < 1:
        raise ValueError("delivery_table must be a non-empty array")
    for i, row in enumerate(dt):
        if not isinstance(row, dict):
            raise ValueError(f"delivery_table[{i}] must be an object")
        for k in ("ticket_id", "committed", "delivered", "notes"):
            if k not in row:
                raise ValueError(f"delivery_table[{i}] missing key {k!r}")
    dr = str(data.get("delivery_rate_summary", "")).strip()
    if not dr:
        raise ValueError("delivery_rate_summary must be non-empty string")

    _validate_narrative_bullets(data.get("went_well"), "went_well", min_items=3, max_items=8)
    _validate_narrative_bullets(
        data.get("did_not_go_well"), "did_not_go_well", min_items=2, max_items=8
    )

    kl = data.get("key_learnings")
    if not isinstance(kl, list) or not (3 <= len(kl) <= 5):
        raise ValueError(f"key_learnings must have between 3 and 5 items, got {kl!r}")
    dec = data.get("decisions")
    if not isinstance(dec, list) or len(dec) < 1:
        raise ValueError("decisions must be a non-empty array")
    actions = data.get("action_items")
    if not isinstance(actions, list) or len(actions) < 1:
        raise ValueError("action_items must be a non-empty array")
    for i, a in enumerate(actions):
        if not isinstance(a, dict):
            raise ValueError(f"action_items[{i}] must be an object")
        for k in ("action", "owner", "due_date"):
            if not str(a.get(k, "")).strip():
                raise ValueError(f"action_items[{i}] needs non-empty {k}")
    rating = data.get("team_health_rating")
    if isinstance(rating, str) and "/" in rating:
        try:
            num = int(rating.split("/")[0].strip())
        except ValueError as e:
            raise ValueError("team_health_rating must be 1-5 or like '4/5'") from e
        if not 1 <= num <= 5:
            raise ValueError("team_health_rating out of range")
    elif isinstance(rating, (int, float)):
        if not 1 <= int(rating) <= 5:
            raise ValueError("team_health_rating must be 1-5")
    else:
        raise ValueError("team_health_rating must be a number 1-5 or string like 4/5")
    if not str(data.get("team_health_note", "")).strip():
        raise ValueError("team_health_note must be non-empty")
    metrics = data.get("metrics")
    if not isinstance(metrics, list) or len(metrics) < 1:
        raise ValueError("metrics must be a non-empty array")
    for i, m in enumerate(metrics):
        if not isinstance(m, dict):
            raise ValueError(
                f"metrics[{i}] must be an object with keys metric and value only, got {type(m).__name__}"
            )
        keys = set(m.keys())
        if keys != {"metric", "value"}:
            raise ValueError(
                f"metrics[{i}] must have exactly keys metric and value (no other keys); got {sorted(keys)}"
            )
        m_raw = m.get("metric", "")
        v_raw = m.get("value", "")
        if m_raw is None or not str(m_raw).strip():
            raise ValueError(f"metrics[{i}].metric is empty or missing")
        if v_raw is None or not str(v_raw).strip():
            raise ValueError(f"metrics[{i}].value is empty or missing")


def _validate_narrative_bullets(items: Any, field: str, *, min_items: int, max_items: int) -> None:
    if not isinstance(items, list):
        raise ValueError(f"{field} must be an array")
    n = len(items)
    if not (min_items <= n <= max_items):
        raise ValueError(f"{field} must have between {min_items} and {max_items} items, got {n}")
    for i, item in enumerate(items):
        if not isinstance(item, str):
            raise ValueError(f"{field}[{i}] must be a string, got {type(item).__name__}")
        t = item.strip()
        sc = _sentence_count(t)
        if sc < 2:
            raise ValueError(
                f"{field}[{i}] must contain at least 2 sentences (sentence_count={sc}); text={t!r}"
            )
        if not _TICKET_ID_RE.search(t):
            raise ValueError(
                f"{field}[{i}] must contain at least one ticket ID like TECH-001 or GOAL-002; text={t!r}"
            )


def _autosize_columns(ws) -> None:
    for col in range(1, ws.max_column + 1):
        max_len = 0
        for row in range(1, ws.max_row + 1):
            v = ws.cell(row=row, column=col).value
            if v is None:
                continue
            max_len = max(max_len, min(len(str(v)), 200))
        ws.column_dimensions[get_column_letter(col)].width = min(max(12, max_len + 2), 70)


def _style_header_row(ws, col_count: int) -> None:
    fill = PatternFill("solid", fgColor="1F4E79")
    font = Font(color="FFFFFF", bold=True)
    for c in range(1, col_count + 1):
        cell = ws.cell(row=1, column=c)
        cell.fill = fill
        cell.font = font
        cell.alignment = Alignment(wrap_text=True, vertical="top")


def _write_retro_workbook(spec: dict[str, Any], out_path: Path, *, sprint_number: int) -> Path:
    wb = Workbook()
    wb.remove(wb.active)

    # --- Sheet 1: Sprint Summary ---
    ws1 = wb.create_sheet(title="Sprint Summary")
    ws1.freeze_panes = "A2"
    headers1 = ["Ticket ID", "Committed", "Delivered", "Notes"]
    ws1.append(headers1)
    _style_header_row(ws1, 4)
    for row in spec["delivery_table"]:
        ws1.append(
            [
                str(row["ticket_id"]).strip(),
                str(row["committed"]).strip(),
                str(row["delivered"]).strip(),
                str(row["notes"]).strip(),
            ]
        )
    ws1.append(["Overall delivery rate", "", "", str(spec["delivery_rate_summary"]).strip()])
    for row in ws1.iter_rows(min_row=1, max_row=ws1.max_row, min_col=1, max_col=4):
        for cell in row:
            cell.alignment = Alignment(wrap_text=True, vertical="top")
    _autosize_columns(ws1)

    # --- Sheet 2: Retrospective Narrative ---
    ws2 = wb.create_sheet(title="Retrospective Narrative")
    ws2.freeze_panes = "A2"
    ws2.column_dimensions["A"].width = 28
    ws2.column_dimensions["B"].width = 55
    ws2.column_dimensions["C"].width = 40

    def add_section(title: str, body: str) -> None:
        r = ws2.max_row + 1
        ws2.cell(row=r, column=1, value=title).font = Font(bold=True)
        ws2.cell(row=r, column=2, value=body).alignment = Alignment(wrap_text=True, vertical="top")

    ws2.cell(row=1, column=1, value="Retrospective Narrative").font = Font(bold=True, size=14)
    ws2.merge_cells(start_row=1, start_column=1, end_row=1, end_column=3)

    add_section("Sprint summary", str(spec["summary"]).strip())

    def bullets_block(items: list[Any], label: str) -> str:
        lines = []
        for j, item in enumerate(items, start=1):
            lines.append(f"{j}. {str(item).strip()}")
        return "\n".join(lines)

    add_section("What went well", bullets_block(list(spec["went_well"]), "went_well"))
    add_section("What did not go well", bullets_block(list(spec["did_not_go_well"]), "did_not_go_well"))
    add_section("Key learnings", bullets_block(list(spec["key_learnings"]), "key_learnings"))
    add_section("Decisions made", bullets_block(list(spec["decisions"]), "decisions"))

    r = ws2.max_row + 1
    ws2.cell(row=r, column=1, value="Action items").font = Font(bold=True)
    r += 1
    for c, h in enumerate(["Action", "Owner", "Due date"], start=1):
        cell = ws2.cell(row=r, column=c, value=h)
        cell.font = Font(bold=True)
        cell.fill = PatternFill("solid", fgColor="1F4E79")
        cell.font = Font(color="FFFFFF", bold=True)
        cell.alignment = Alignment(wrap_text=True, vertical="top")
    for a in spec["action_items"]:
        r += 1
        ws2.cell(row=r, column=1, value=str(a["action"]).strip()).alignment = Alignment(
            wrap_text=True, vertical="top"
        )
        ws2.cell(row=r, column=2, value=str(a["owner"]).strip()).alignment = Alignment(
            wrap_text=True, vertical="top"
        )
        ws2.cell(row=r, column=3, value=str(a["due_date"]).strip()).alignment = Alignment(
            wrap_text=True, vertical="top"
        )

    r += 1
    ws2.cell(row=r, column=1, value="Team health rating").font = Font(bold=True)
    ws2.cell(row=r, column=2, value=spec["team_health_rating"])
    ws2.cell(row=r, column=3, value=str(spec["team_health_note"]).strip()).alignment = Alignment(
        wrap_text=True, vertical="top"
    )

    for row in ws2.iter_rows(min_row=2, max_row=ws2.max_row, min_col=1, max_col=3):
        for cell in row:
            if cell.alignment is None or cell.alignment.horizontal is None:
                cell.alignment = Alignment(wrap_text=True, vertical="top")

    # --- Sheet 3: Metrics This Sprint ---
    ws3 = wb.create_sheet(title="Metrics This Sprint")
    ws3.freeze_panes = "A2"
    ws3.append(["Metric", "Value"])
    _style_header_row(ws3, 2)
    for m in spec["metrics"]:
        ws3.append([str(m["metric"]).strip(), str(m["value"]).strip()])
    for row in ws3.iter_rows(min_row=2, max_row=ws3.max_row, min_col=1, max_col=2):
        for cell in row:
            cell.alignment = Alignment(wrap_text=True, vertical="top")
    _autosize_columns(ws3)

    out_path.parent.mkdir(parents=True, exist_ok=True)
    tmp = out_path.with_suffix(".tmp.xlsx")
    wb.save(tmp)
    try:
        os.replace(tmp, out_path)
        return out_path
    except PermissionError:
        fb = out_path.with_name(f"sprint{sprint_number}_retro_ready.xlsx")
        shutil.move(str(tmp), str(fb))
        return fb


def generate_sprint_retro(sprint_number: int) -> Path:
    md_full = SOURCE_MARKDOWN.read_text(encoding="utf-8")
    blueprint = _extract_retro_blueprint(md_full, sprint_number)
    out_path = _retro_out_path(sprint_number)

    user = (
        f"Return ONLY valid JSON (no markdown, no code fences) for Sprint {sprint_number} retrospective Excel content. "
        "Follow the SYSTEM prompt. Use the blueprint as source of truth for tickets, dates, names, and outcomes — "
        "rewrite narrative text in Shristi's voice; do not copy the blueprint verbatim.\n\n"
        "Required keys:\n"
        "- summary: string, 2-6 sentences, overall what happened this sprint\n"
        "- delivery_table: array of objects, each with ticket_id, committed, delivered, notes "
        '(Yes/No or short text for committed/delivered; notes like the blueprint)\n'
        "- delivery_rate_summary: string, one line summarising overall delivery (e.g. '8 out of 8 tickets. 100%.')\n"
        "- went_well: array of 3 to 8 strings. Each string must be at least 2 full sentences and include "
        "at least one ticket ID (e.g. TECH-001, GOAL-SPIKE-001).\n"
        "- did_not_go_well: array of 2 to 8 strings. Each string must be at least 2 full sentences and include "
        "at least one ticket ID.\n"
        "- key_learnings: array of exactly 3 to 5 strings, actionable\n"
        "- decisions: array of strings, decisions made this sprint\n"
        "- action_items: array of objects with action, owner, due_date (specific dates or day/month)\n"
        "- team_health_rating: integer 1-5 OR string like '4/5'\n"
        "- team_health_note: string explaining the rating\n"
        "- metrics: non-empty array. Each item MUST be an object with ONLY these keys: metric, value. "
        "Both must be non-empty strings (value must include concrete numbers or facts plus brief context). "
        "Do not use any other keys. Do not return strings for metrics entries.\n\n"
        "Blueprint:\n-----\n"
        f"{blueprint}\n-----\n"
    )

    print(f"Sprint {sprint_number}: calling API...")
    client = _client()
    spec = _call_retro_json(client, user=user)
    saved = _write_retro_workbook(spec, out_path, sprint_number=sprint_number)
    if saved != out_path:
        print(
            f"Sprint {sprint_number}: saved to {saved} (could not overwrite {out_path} — file may be open). "
            f"Close sprint{sprint_number}_retro.xlsx and rerun."
        )
    else:
        print(f"Sprint {sprint_number}: wrote {saved}")
    return saved


def main(argv: list[str] | None = None) -> int:
    parser = argparse.ArgumentParser(description="Generate Nutrivana sprint retrospective Excel files.")
    parser.add_argument(
        "--all",
        action="store_true",
        help="Generate retrospectives for sprints 1 through 9.",
    )
    parser.add_argument(
        "sprints",
        nargs="*",
        type=int,
        metavar="N",
        help="Sprint numbers 1-9 (default: 1 only). Ignored if --all is set.",
    )
    args = parser.parse_args(argv)
    if args.all:
        sprints = list(range(1, 10))
    elif args.sprints:
        sprints = args.sprints
    else:
        sprints = [1]

    for n in sprints:
        if n < 1 or n > 9:
            print(f"[error] Sprint number must be 1-9, got {n}", file=sys.stderr)
            return 1

    print(f"Model: {MODEL}")
    try:
        for n in sprints:
            generate_sprint_retro(n)
    except GenerationError as e:
        print(str(e), file=sys.stderr)
        return 1
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
